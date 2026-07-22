import os
import tempfile
import time
from decimal import Decimal
from django.test import TestCase
from django.contrib.auth import get_user_model
from dies.models import Die, RoundDie, FlatDie
from machines.models import Rack
from dies.services.import_service import ImportService
import_dies = ImportService.import_dies
from search.meili import client as meili_client, INDEX_NAME

def wait_for_meili_tasks():
    try:
        for _ in range(50):
            tasks = meili_client.get_tasks({'statuses': ['enqueued', 'processing'], 'indexUids': [INDEX_NAME]})
            if not tasks.results:
                break
            time.sleep(0.1)
    except Exception:
        pass

User = get_user_model()

class BulkImportTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username='importer',
            password='password123',
            role='ADMIN'
        )
        self.rack_a = Rack.objects.create(name="Rack A", row_count=4, column_count=3)
        self.rack_b = Rack.objects.create(name="Rack B", row_count=4, column_count=3)
        self.rack_c = Rack.objects.create(name="Rack C", row_count=4, column_count=3)
        try:
            meili_client.index(INDEX_NAME).delete()
        except Exception:
            pass
        try:
            meili_client.create_index(INDEX_NAME, {'primaryKey': 'id'})
        except Exception:
            pass
        try:
            index = meili_client.index(INDEX_NAME)
            task = index.update_settings({
                'searchableAttributes': ['die_id', 'casing', 'status', 'rack', 'shelf_number', 'set', 'machine', 'size', 'width', 'thickness'],
                'filterableAttributes': ['die_type', 'status', 'casing', 'rack', 'shelf_number', 'size', 'width', 'thickness', 'machine'],
                'sortableAttributes':   ['die_id'],
            })
            uid = task.get('taskUid') or task.get('task_uid')
            meili_client.wait_for_task(uid, timeout_in_ms=5000)
        except Exception:
            pass

        try:
            task = meili_client.index(INDEX_NAME).delete_all_documents()
            uid = task.get('taskUid') or task.get('task_uid')
            meili_client.wait_for_task(uid, timeout_in_ms=5000)
        except Exception:
            time.sleep(0.5)

    def write_temp_csv(self, content):
        temp_file = tempfile.NamedTemporaryFile(suffix='.csv', delete=False, mode='w', encoding='utf-8')
        temp_file.write(content)
        temp_file.close()
        return temp_file.name

    def test_import_new_round_dies(self):
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-1,ROUND,25x10,AVAILABLE,Rack A,1,remark1,2.5,2.5
R-IMP-2,ROUND,25x10,RUNNING,Rack A,2,remark2,3.0,3.0
R-IMP-3,ROUND,25x10,CLEANING,Rack B,1,remark3,3.5,3.5
R-IMP-4,ROUND,25x10,POLISHING,Rack B,2,remark4,4.0,4.0
R-IMP-5,ROUND,25x10,DAMAGED,Rack C,1,remark5,4.5,4.5
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 5)
            self.assertEqual(res['updated'], 0)
            self.assertEqual(len(res['errors']), 0)

            # Check DB count
            self.assertEqual(Die.objects.filter(die_id__startswith='R-IMP-').count(), 5)
            self.assertEqual(RoundDie.objects.count(), 5)

            # Verify details
            rd = Die.objects.get(die_id='R-IMP-1')
            self.assertEqual(rd.status, 'AVAILABLE')
            self.assertEqual(rd.rounddie.current_size, Decimal('2.500'))

            # Wait for Meilisearch sync and check
            wait_for_meili_tasks()
            index = meili_client.index(INDEX_NAME)
            try:
                print("INDEX DOCUMENTS:", [d.__dict__ for d in index.get_documents().results])
            except Exception as e:
                print("FAILED TO GET DOCUMENTS:", e)
            doc = index.get_document(str(rd.id))
            self.assertEqual(doc.id, str(rd.id))
            self.assertEqual(doc.status, "AVAILABLE")
        finally:
            os.remove(filepath)

    def test_import_existing_and_new_dies(self):
        die1 = Die.objects.create(die_id='R-IMP-1', die_type='ROUND', casing='25x10', status='AVAILABLE')
        RoundDie.objects.create(die=die1, punched_size=Decimal('2.5'), current_size=Decimal('2.5'))

        die2 = Die.objects.create(die_id='R-IMP-2', die_type='ROUND', casing='25x10', status='RUNNING')
        RoundDie.objects.create(die=die2, punched_size=Decimal('3.0'), current_size=Decimal('3.0'))

        die3 = Die.objects.create(die_id='R-IMP-3', die_type='ROUND', casing='25x10', status='CLEANING')
        RoundDie.objects.create(die=die3, punched_size=Decimal('3.5'), current_size=Decimal('3.5'))

        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-1,ROUND,25x10,RUNNING,Rack A,1,remark1,2.5,2.5
R-IMP-2,ROUND,25x10,AVAILABLE,Rack A,2,remark2,3.0,3.0
R-IMP-3,ROUND,25x10,AVAILABLE,Rack B,1,remark3,3.5,3.5
R-IMP-6,ROUND,25x10,AVAILABLE,Rack C,1,,5.0,5.0
R-IMP-7,ROUND,25x10,AVAILABLE,Rack C,2,,5.5,5.5
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 2) # R-IMP-6, R-IMP-7
            self.assertEqual(res['updated'], 3) # R-IMP-1, R-IMP-2, R-IMP-3
            self.assertEqual(len(res['errors']), 0)
        finally:
            os.remove(filepath)

    def test_import_missing_die_id_header(self):
        content = """die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
ROUND,25x10,AVAILABLE,Rack A,1,remark1,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 0)
            self.assertEqual(len(res['errors']), 1)
            self.assertIn("Missing 'die_id'", res['errors'][0]['error'])
        finally:
            os.remove(filepath)

    def test_import_invalid_status_skips_row_but_others_succeed(self):
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-8,ROUND,25x10,INVALID_STATUS,Rack A,1,remark1,2.5,2.5
R-IMP-9,ROUND,25x10,AVAILABLE,Rack A,1,remark9,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 1)
            self.assertEqual(len(res['errors']), 1)
            self.assertEqual(res['errors'][0]['row'], 2)
            self.assertIn("Invalid status", res['errors'][0]['error'])

            # Check only R-IMP-9 created
            self.assertTrue(Die.objects.filter(die_id='R-IMP-9').exists())
            self.assertFalse(Die.objects.filter(die_id='R-IMP-8').exists())
        finally:
            os.remove(filepath)

    def test_import_idempotent_double_import(self):
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-10,ROUND,25x10,AVAILABLE,Rack A,1,,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            # First import -> created
            res1 = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res1['created'], 1)
            self.assertEqual(res1['updated'], 0)

            # Second import -> updated (no duplicates created)
            res2 = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res2['created'], 0)
            self.assertEqual(res2['updated'], 1)
        finally:
            os.remove(filepath)

    def test_import_flat_dies(self):
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_width,current_width,punched_thickness,current_thickness,radius
F-IMP-1,FLAT,30x15,AVAILABLE,Rack B,1,,5.5,5.5,15.0,15.0,1.0
F-IMP-2,FLAT,30x15,RUNNING,Rack B,2,,6.0,6.0,16.0,16.0,1.5
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 2)
            self.assertEqual(len(res['errors']), 0)

            # Verify flat die creation
            die1 = Die.objects.get(die_id='F-IMP-1')
            self.assertEqual(die1.die_type, 'FLAT')
            self.assertEqual(die1.flatdie.current_width, Decimal('5.500'))
            self.assertEqual(die1.flatdie.current_thickness, Decimal('15.000'))

            # Wait for Meilisearch sync and check
            wait_for_meili_tasks()
            index = meili_client.index(INDEX_NAME)
            doc = index.get_document(str(die1.id))
            self.assertEqual(doc.id, str(die1.id))
            self.assertEqual(getattr(doc, 'width', None), 5.5)
            self.assertEqual(getattr(doc, 'thickness', None), 15.0)
        finally:
            os.remove(filepath)

    def test_import_with_set_name(self):
        from machines.models import MachineCategory, Machine, Set

        cat = MachineCategory.objects.create(name="Cat Import")
        mach1 = Machine.objects.create(name="Mach A", category=cat)
        mach2 = Machine.objects.create(name="Mach B", category=cat)

        set_unique = Set.objects.create(name="UniqueSet", machine=mach1)
        set_dup1 = Set.objects.create(name="DupSet", machine=mach1)
        set_dup2 = Set.objects.create(name="DupSet", machine=mach2)

        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,set_name,machine_name,punched_size,current_size
R-IMP-101,ROUND,25x10,AVAILABLE,Rack A,1,,UniqueSet,,2.5,2.5
R-IMP-102,ROUND,25x10,AVAILABLE,Rack A,2,,DupSet,Mach B,3.0,3.0
"""
        filepath = self.write_temp_csv(content)
        try:
            res = import_dies(filepath, '.csv', self.user)
            self.assertEqual(res['created'], 2)
            self.assertEqual(len(res['errors']), 0)

            d1 = Die.objects.get(die_id='R-IMP-101')
            self.assertEqual(d1.current_set, set_unique)

            d2 = Die.objects.get(die_id='R-IMP-102')
            self.assertEqual(d2.current_set, set_dup2)
        finally:
            os.remove(filepath)

    def test_import_dry_run_mode(self):
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-DRY,ROUND,25x10,AVAILABLE,Rack A,1,remark1,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            # Run dry run
            res = import_dies(filepath, '.csv', self.user, dry_run=True)
            self.assertEqual(res['created'], 1)
            self.assertEqual(res['updated'], 0)
            self.assertTrue(res['dry_run'])

            # Verify record does NOT exist in DB
            self.assertFalse(Die.objects.filter(die_id='R-IMP-DRY').exists())
        finally:
            os.remove(filepath)

    def test_import_view_dry_run_endpoint(self):
        from django.urls import reverse
        from rest_framework import status
        from rest_framework_simplejwt.tokens import AccessToken
        from users.models import UserSession
        from rest_framework.test import APIClient
        import hashlib

        token = str(AccessToken.for_user(self.user))
        UserSession.objects.create(
            user=self.user,
            token_hash=hashlib.sha256(token.encode('utf-8')).hexdigest()
        )

        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        url = reverse('import-dies') + '?dry_run=true'
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-IMP-VIEW-DRY,ROUND,25x10,AVAILABLE,Rack A,1,remark1,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            with open(filepath, 'rb') as f:
                response = client.post(url, {'file': f}, format='multipart')
            self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)

            # Verify database was not changed
            self.assertFalse(Die.objects.filter(die_id='R-IMP-VIEW-DRY').exists())
        finally:
            os.remove(filepath)

    def test_import_template_download_endpoint(self):
        from django.urls import reverse
        from rest_framework import status
        from rest_framework_simplejwt.tokens import AccessToken
        from users.models import UserSession
        from rest_framework.test import APIClient
        import hashlib
        import openpyxl

        token = str(AccessToken.for_user(self.user))
        UserSession.objects.create(
            user=self.user,
            token_hash=hashlib.sha256(token.encode('utf-8')).hexdigest()
        )

        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        url = reverse('import-template')
        response = client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('dms_import_template.xlsx', response['Content-Disposition'])

        # Verify it is a valid openpyxl workbook
        import io
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertIn("Round Die", wb.sheetnames)
        self.assertIn("Flat Die", wb.sheetnames)
        self.assertIn("Field Reference", wb.sheetnames)

    def test_import_log_persistence(self):
        from dies.models import ImportLog
        from django.urls import reverse
        from rest_framework import status
        from rest_framework_simplejwt.tokens import AccessToken
        from users.models import UserSession
        from rest_framework.test import APIClient
        import hashlib

        token = str(AccessToken.for_user(self.user))
        UserSession.objects.create(
            user=self.user,
            token_hash=hashlib.sha256(token.encode('utf-8')).hexdigest()
        )

        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')

        url = reverse('import-dies')
        content = """die_id,die_type,casing,status,rack,shelf_number,remarks,punched_size,current_size
R-LOG-1,ROUND,25x10,AVAILABLE,Rack A,1,remark1,2.5,2.5
"""
        filepath = self.write_temp_csv(content)
        try:
            with open(filepath, 'rb') as f:
                response = client.post(url, {'file': f}, format='multipart')
            self.assertEqual(response.status_code, status.HTTP_202_ACCEPTED)

            # Check ImportLog was created
            log_entry = ImportLog.objects.filter(filename=os.path.basename(filepath)).first()
            self.assertIsNotNone(log_entry)
            self.assertEqual(log_entry.created_count, 1)
            self.assertEqual(log_entry.imported_by, self.user)

            # Verify list import logs endpoint
            logs_url = reverse('import-logs')
            logs_res = client.get(logs_url)
            self.assertEqual(logs_res.status_code, status.HTTP_200_OK)
            self.assertTrue(len(logs_res.data['results']) >= 1)
        finally:
            os.remove(filepath)
