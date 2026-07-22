import io
import openpyxl
from django.db import transaction
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from dies.models import Die, ImportLog, MaintenanceLog, DieTolerance, WearAlert
from dies.serializers import (
    DieListSerializer, DieDetailSerializer, DieCreateSerializer, 
    serialize_die_list_fast, ImportLogSerializer, MaintenanceLogSerializer,
    DieToleranceSerializer, WearAlertSerializer
)
from users.permissions import IsAdminOrRoot, IsAdminOrRootOrOperatorRelocate, IsAdminOrRootOnly
from search.meili import client as meili_client, INDEX_NAME

class DieViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing die inventory.
    
    Supports full CRUD operations for dies with filtering and pagination.
    
    - **Create** (POST): Add new die (ROUND or FLAT type)
    - **Read** (GET): List all dies with optional filters, or retrieve single die
    - **Update** (PUT/PATCH): Modify die properties
    - **Delete** (DELETE): Remove die from inventory
    
    **Filtering Parameters (GET only):**
    - `die_type`: Filter by die type (ROUND or FLAT)
    - `status`: Filter by status (ACTIVE, INACTIVE, AVAILABLE, etc.)
    - `casing`: Filter by casing material (STEEL, CARBIDE, etc.)
    - `location`: Filter by storage location (partial match, case-insensitive)
    - `size_min`, `size_max`: Range filter for ROUND die current_size
    - `width_min`, `width_max`: Range filter for FLAT die current_width
    - `thick_min`, `thick_max`: Range filter for FLAT die current_thickness
    - `page`: Pagination (100 items per page)
    
    **Examples:**
    - List ROUND dies with size 5.0-5.5: `/api/dies/?die_type=ROUND&size_min=5.0&size_max=5.5`
    - List ACTIVE dies in RACK-A: `/api/dies/?status=ACTIVE&location=RACK-A`
    - List FLAT dies by thickness: `/api/dies/?die_type=FLAT&thick_min=2.0&thick_max=2.5`
    
    **Permissions:** ADMIN, ROOT or OPERATOR (for relocation only)
    
    **Authentication:** Required (JWT Bearer token)
    """
    queryset = Die.objects.select_related('rounddie', 'flatdie', 'current_set__machine', 'rack').prefetch_related('wear_alerts', 'history')
    lookup_field = 'die_id'
    lookup_value_regex = r'(?:(?!/wear-prediction(?:/|$)|/maintenance-logs(?:/|$)|/recut(?:/|$))[^?#])+'
    permission_classes = [IsAdminOrRootOrOperatorRelocate]

    def get_serializer_class(self):
        if self.action == 'list':
            return DieListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return DieCreateSerializer
        return DieDetailSerializer

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # 1. OPTIMISTIC CONCURRENCY CONTROL (OCC)
        client_version = request.data.get('version')
        if client_version is not None:
            try:
                client_version = int(client_version)
            except ValueError:
                return Response({"detail": "Invalid version format."}, status=status.HTTP_400_BAD_REQUEST)
                
            if instance.version != client_version:
                return Response({
                    "detail": "Concurrency Conflict: This record has been modified by another user. Please refresh and try again.",
                    "code": "concurrency_conflict"
                }, status=status.HTTP_409_CONFLICT)
                
        # 2. PESSIMISTIC LOCKING
        locked_instance = Die.objects.select_for_update().get(die_id=instance.die_id)
        
        serializer = self.get_serializer(locked_instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        
        locked_instance.version += 1
        locked_instance.save(update_fields=['version'])
        
        self.perform_update(serializer)
        return Response(serializer.data)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            data = serialize_die_list_fast(page)
            return self.get_paginated_response(data)
        data = serialize_die_list_fast(queryset)
        return Response(data)

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Query parameters for filtering
        die_type = self.request.query_params.get('die_type')
        status_val = self.request.query_params.get('status')
        casing = self.request.query_params.get('casing')
        rack_id = self.request.query_params.get('rack_id')
        shelf_number = self.request.query_params.get('shelf_number')
        
        # Range queries for ROUND size and FLAT width/thickness
        size_min = self.request.query_params.get('size_min')
        size_max = self.request.query_params.get('size_max')
        
        width_min = self.request.query_params.get('width_min')
        width_max = self.request.query_params.get('width_max')
        thick_min = self.request.query_params.get('thick_min')
        thick_max = self.request.query_params.get('thick_max')
        
        if die_type:
            queryset = queryset.filter(die_type=die_type)
        if status_val:
            queryset = queryset.filter(status=status_val)
        if casing:
            queryset = queryset.filter(casing=casing)
        if rack_id:
            queryset = queryset.filter(rack_id=rack_id)
        if shelf_number:
            queryset = queryset.filter(shelf_number=shelf_number)
            
        if size_min:
            queryset = queryset.filter(rounddie__current_size__gte=size_min)
        if size_max:
            queryset = queryset.filter(rounddie__current_size__lte=size_max)
            
        if width_min:
            queryset = queryset.filter(flatdie__current_width__gte=width_min)
        if width_max:
            queryset = queryset.filter(flatdie__current_width__lte=width_max)
        if thick_min:
            queryset = queryset.filter(flatdie__current_thickness__gte=thick_min)
        if thick_max:
            queryset = queryset.filter(flatdie__current_thickness__lte=thick_max)
            
        return queryset

    @action(detail=True, methods=['get', 'post'], url_path='maintenance-logs')
    def maintenance_logs(self, request, die_id=None):
        die = self.get_object()
        if request.method == 'GET':
            logs = die.maintenance_logs.select_related('created_by').all()
            serializer = MaintenanceLogSerializer(logs, many=True)
            return Response(serializer.data)
        elif request.method == 'POST':
            serializer = MaintenanceLogSerializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            serializer.save(die=die, created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='recut')
    @transaction.atomic
    def recut(self, request, die_id=None):
        die = self.get_object()
        from dies.services.recut_service import RecutService
        try:
            RecutService.recut_die(die, request.user, request.data)
        except PermissionError as e:
            return Response({"detail": str(e)}, status=status.HTTP_403_FORBIDDEN)
        except ValueError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({"detail": "Die recut successfully."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='wear-prediction')
    def wear_prediction(self, request, die_id=None):
        user = request.user
        is_authorized = (
            user and user.is_authenticated and (
                user.role in ['ROOT', 'ADMIN'] or
                user.is_superuser
            )
        )
        if not is_authorized:
            return Response({"detail": "You do not have permission to view wear prediction details."}, status=status.HTTP_403_FORBIDDEN)
        die = self.get_object()
        from django.core.cache import cache
        cache_key = f"die_wear_prediction_{die.id}"
        analysis = cache.get(cache_key)
        if analysis is None:
            from dies.services.wear_prediction_service import WearPredictionService
            try:
                analysis = WearPredictionService.predict_die(die)
                cache.set(cache_key, analysis, timeout=86400)
            except ValueError as e:
                return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(analysis)




import tempfile
import os
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView
from rest_framework import serializers
from drf_spectacular.utils import OpenApiResponse, extend_schema, inline_serializer
from dies.services.import_service import ImportService

class ImportDiesView(APIView):
    """
    Bulk import dies from CSV or XLSX file.
    
    **Description:**
    Upload a spreadsheet file to create multiple dies at once.
    Supports CSV and XLSX formats with columns:
    - die_id, die_type (ROUND|FLAT), casing, status, location, current_set
    - For ROUND: punched_size, current_size
    - For FLAT: punched_width, current_width, punched_thickness, current_thickness, radius
    
    **Request:**
    - Method: POST
    - Content-Type: multipart/form-data
    - Field: `file` (CSV or XLSX file)
    
    **Response:**
    - `201 Created`: Import completed with summary
    - `400 Bad Request`: Invalid file format or missing file
    
    **Example Response:**
    ```json
    {
      "imported": 42,
      "skipped": 3,
      "errors": [
        {"row": 5, "error": "Invalid die_type: INVALID"}
      ]
    }
    ```
    
    **Permissions:** ADMIN or ROOT user only
    **Authentication:** Required (JWT Bearer token)
    """
    permission_classes = [IsAdminOrRoot]
    parser_classes = [MultiPartParser]

    @extend_schema(
        request={
            'multipart/form-data': inline_serializer(
                name='ImportDiesRequest',
                fields={'file': serializers.FileField()},
            )
        },
        responses={
            200: OpenApiResponse(description='Import completed with summary'),
            400: OpenApiResponse(description='Invalid file upload'),
        },
    )
    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)
        
        name, ext = os.path.splitext(file_obj.name)
        if ext.lower() not in ['.csv', '.xlsx']:
            return Response({"error": "Unsupported file format. Only CSV and XLSX are supported."}, status=status.HTTP_400_BAD_REQUEST)

        dry_run = request.query_params.get('dry_run', '').lower() == 'true'

        from django.conf import settings
        tmp_dir = os.path.join(settings.BASE_DIR, 'tmp')
        os.makedirs(tmp_dir, exist_ok=True)

        with tempfile.NamedTemporaryFile(suffix=ext, delete=False, dir=tmp_dir) as temp_file:
            for chunk in file_obj.chunks():
                temp_file.write(chunk)
            temp_path = temp_file.name

        try:
            import json
            from django.core.cache import cache
            from dies.tasks import import_dies_task

            initial_status = {
                "status": "importing",
                "progress": 0,
                "total": 100,
                "filename": file_obj.name,
                "dry_run": dry_run
            }
            cache.set('import_status', json.dumps(initial_status), timeout=3600)

            import_dies_task.delay(temp_path, ext, request.user.username, file_obj.name, dry_run=dry_run)
            
            return Response({
                "detail": "Import started in background",
                "status": "importing"
            }, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return Response({"error": f"Failed to start import task: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(transaction.non_atomic_requests, name='dispatch')
class ImportTemplateView(APIView):
    """
    Download Excel import templates.
    """
    permission_classes = [IsAdminOrRoot]

    @extend_schema(
        responses={200: OpenApiResponse(description="Excel spreadsheet binary template file")},
        description="Download an Excel template containing tabs for Round Die and Flat Die structure, and validation rules."
    )
    def get(self, request, *args, **kwargs):
        wb = openpyxl.Workbook()
        # Sheet 1: Round Die
        ws1 = wb.active
        ws1.title = "Round Die"
        headers_round = [
            "die_id", "die_type", "casing", "status", "rack", "shelf_number", "remarks", 
            "current_set", "machine_name", "punched_size", "current_size"
        ]
        ws1.append(headers_round)
        example_round = [
            "R-999", "ROUND", "25x10", "AVAILABLE", "A", 3, "Example round die",
            "Set A", "Machine A", 12.345, 12.345
        ]
        ws1.append(example_round)

        # Sheet 2: Flat Die
        ws2 = wb.create_sheet(title="Flat Die")
        headers_flat = [
            "die_id", "die_type", "casing", "status", "rack", "shelf_number", "remarks",
            "current_set", "machine_name", "punched_width", "current_width",
            "punched_thickness", "current_thickness", "radius"
        ]
        ws2.append(headers_flat)
        example_flat = [
            "F-999", "FLAT", "30x15", "AVAILABLE", "B", 1, "Example flat die",
            "Set B", "Machine B", 50.123, 50.123, 15.456, 15.456, 1.500
        ]
        ws2.append(example_flat)

        # Sheet 3: Field Reference
        ws3 = wb.create_sheet(title="Field Reference")
        ws3.append(["Column Name", "Required/Optional", "Die Type", "Description / Valid Values"])
        reference_rows = [
            ["die_id", "Required", "Both", "Unique identifier for the die (e.g., R-101, F-202)"],
            ["die_type", "Required", "Both", "Must be 'ROUND' or 'FLAT'"],
            ["casing", "Required", "Both", "Physical casing dimensions (e.g., 25x10, 30x15)"],
            ["status", "Required", "Both", "Must be one of: AVAILABLE, RUNNING, CLEANING, DAMAGED, POLISHING, MEASURING, INACTIVE"],
            ["rack", "Optional", "Both", "Rack name (e.g., A, B, C) - must match existing rack"],
            ["shelf_number", "Optional", "Both", "Shelf number within rack (must be within rack dimensions)"],
            ["remarks", "Optional", "Both", "Free-form text comments"],
            ["current_set", "Optional", "Both", "The Set name or Set ID the die belongs to"],
            ["machine_name", "Optional", "Both", "Machine name to resolve duplicate set names"],
            ["punched_size", "Required", "ROUND", "Punched extrusion size currently marked on die in mm (decimal)"],
            ["current_size", "Required", "ROUND", "Current measured extrusion size in mm (decimal)"],
            ["punched_width", "Required", "FLAT", "Punched flat die width currently marked on die in mm (decimal)"],
            ["current_width", "Required", "FLAT", "Current flat die width in mm (decimal)"],
            ["punched_thickness", "Required", "FLAT", "Punched flat die thickness currently marked on die in mm (decimal)"],
            ["current_thickness", "Required", "FLAT", "Current flat die thickness in mm (decimal)"],
            ["radius", "Required", "FLAT", "Corner/bearing radius in mm (decimal)"],
        ]
        for r in reference_rows:
            ws3.append(r)

        # Autofit column widths for readability
        for sheet in [ws1, ws2, ws3]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dms_import_template.xlsx"'
        return response


@method_decorator(transaction.non_atomic_requests, name='dispatch')
class ImportLogsView(APIView):
    """
    Get paginated history of imports.
    """
    permission_classes = [IsAdminOrRoot]

    @extend_schema(
        responses={200: ImportLogSerializer(many=True)},
        description="Retrieve history of bulk imports."
    )
    def get(self, request, *args, **kwargs):
        from rest_framework.pagination import PageNumberPagination
        paginator = PageNumberPagination()
        paginator.page_size = 50
        logs = ImportLog.objects.select_related('imported_by').all()
        result_page = paginator.paginate_queryset(logs, request, view=self)
        serializer = ImportLogSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)


class DieToleranceViewSet(viewsets.ModelViewSet):
    queryset = DieTolerance.objects.all()
    serializer_class = DieToleranceSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrRoot]


class WearAlertViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = WearAlert.objects.select_related('die').all()
    serializer_class = WearAlertSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrRootOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        is_resolved = self.request.query_params.get('is_resolved')
        if is_resolved is not None:
            qs = qs.filter(is_resolved=is_resolved.lower() == 'true')
        die_id = self.request.query_params.get('die_id')
        if die_id is not None:
            qs = qs.filter(die__die_id=die_id)
        return qs
